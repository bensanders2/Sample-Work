"""
change_audit_pipeline.py
------------------------
Automated IT Change Management Audit Pipeline

Connects to a Snowflake data warehouse to extract change records,
applies stratified sampling for minor change audits, computes
compliance metrics, exports styled Excel reports, and sends
summary notifications via email.

Packages: pandas, snowflake-connector-python, openpyxl, python-dotenv, smtplib
"""

# ===========================================================================
# IMPORTS
# ===========================================================================

import os                          # Used to read environment variables and manage file paths
import smtplib                     # Built-in Python library for sending emails over SMTP
import logging                     # Built-in library for writing timestamped log messages to the console
from datetime import datetime, timedelta   # datetime: for timestamps/filenames; timedelta: for date math (not used directly but useful for extensions)
from email.mime.multipart import MIMEMultipart  # Constructs multi-part email messages (body + attachments together)
from email.mime.text import MIMEText            # Wraps plain text or HTML content as an email body part
from email.mime.base import MIMEBase            # Base class for adding binary attachments (e.g., Excel files) to email
from email import encoders                      # Encodes binary attachments in base64 so they survive email transmission

import pandas as pd                         # Core data manipulation library — used for all DataFrame operations
import snowflake.connector                  # Official Snowflake Python connector for executing SQL and fetching results
from dotenv import load_dotenv              # Loads key=value pairs from a .env file into os.getenv() — keeps credentials out of source code
from openpyxl import load_workbook          # Opens an existing .xlsx file so we can apply styles after pandas writes it
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # openpyxl style objects for cell formatting
from openpyxl.utils import get_column_letter  # Converts a column number (e.g., 3) to its Excel letter (e.g., "C")


# ===========================================================================
# CONFIGURATION
# ===========================================================================

load_dotenv()  # Reads the .env file in the project root and loads all variables into the environment

# Set up logging so every log.info() call prints a timestamped message like:
# 2025-06-01 09:14:22 [INFO] Connecting to Snowflake...
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)  # Creates a logger scoped to this module's name

# Dictionary of Snowflake connection parameters — values pulled from .env so credentials are never hardcoded
SNOWFLAKE_CREDS = {
    "account":   os.getenv("SF_ACCOUNT"),             # Snowflake account identifier (e.g., "abc12345.us-east-1")
    "user":      os.getenv("SF_USER"),                 # Service account username
    "password":  os.getenv("SF_PASSWORD"),             # Password for the service account
    "warehouse": os.getenv("SF_WAREHOUSE", "COMPUTE_WH"),         # Virtual warehouse to use; falls back to "COMPUTE_WH" if not set
    "database":  os.getenv("SF_DATABASE"),             # Target Snowflake database
    "schema":    os.getenv("SF_SCHEMA", "SERVICE_MANAGEMENT"),    # Default schema within the database
    "role":      os.getenv("SF_ROLE", "ANALYST_READ"),            # Snowflake role — read-only to match production access constraints
}

SMTP_HOST  = os.getenv("SMTP_HOST", "smtp.example.gov")   # Mail server hostname (defaults to a placeholder if not set)
SMTP_PORT  = int(os.getenv("SMTP_PORT", 587))             # SMTP port; 587 is standard for TLS/STARTTLS
SENDER     = os.getenv("EMAIL_SENDER")                    # "From" address for outbound emails
RECIPIENTS = os.getenv("EMAIL_RECIPIENTS", "").split(",") # Comma-separated list of recipients, split into a Python list

AUDIT_SAMPLE_RATE = 0.10   # Pull 10% of standard changes per support group for the monthly audit review
LOOKBACK_DAYS     = 30     # How many calendar days back to pull change records
OUTPUT_DIR        = "reports"              # Local folder where Excel reports will be saved
os.makedirs(OUTPUT_DIR, exist_ok=True)    # Create the folder if it doesn't exist yet; no error if it already does


# ===========================================================================
# SNOWFLAKE HELPERS
# ===========================================================================

def get_connection():
    """Open and return a Snowflake connection using the credentials defined above."""
    log.info("Connecting to Snowflake...")
    # ** unpacks the SNOWFLAKE_CREDS dictionary as keyword arguments to the connector
    return snowflake.connector.connect(**SNOWFLAKE_CREDS)


def run_query(conn, sql: str, params: dict | None = None) -> pd.DataFrame:
    """
    Execute a parameterized SQL query and return the results as a pandas DataFrame.

    Using a cursor (rather than a raw connection) is best practice — it isolates
    each query's state and ensures the cursor is always closed even if an error occurs.
    """
    cursor = conn.cursor()   # Open a cursor object to send SQL commands through the connection
    try:
        cursor.execute(sql, params or {})   # Run the SQL; params fills in %(key)s placeholders safely (prevents SQL injection)
        cols = [desc[0] for desc in cursor.description]   # cursor.description is a list of tuples; index 0 of each is the column name
        rows = cursor.fetchall()                           # Retrieve all result rows as a list of tuples
        return pd.DataFrame(rows, columns=cols)            # Combine column names + rows into a DataFrame
    finally:
        cursor.close()   # Always close the cursor, even if the query raised an exception


# ===========================================================================
# DATA EXTRACTION
# ===========================================================================

# SQL query stored as a module-level constant — keeps it readable and separate from logic
# %(lookback)s is a named parameter placeholder filled in at runtime by run_query()
CHANGE_QUERY = """
    SELECT
        c.CHANGE_ID,              -- Unique identifier for each change request
        c.CHANGE_TYPE,            -- Standard, Normal, or Emergency
        c.STATUS,                 -- Current workflow status (e.g., Closed, Implemented)
        c.RISK_LEVEL,             -- Risk tier assigned during approval
        c.COORDINATOR,            -- Employee ID of the change coordinator
        c.SUPPORT_GROUP,          -- Team responsible for implementing the change
        c.SCHEDULED_START,        -- Planned start timestamp
        c.SCHEDULED_END,          -- Planned end timestamp
        c.ACTUAL_START,           -- Actual start timestamp recorded at implementation
        c.ACTUAL_END,             -- Actual end timestamp recorded at closure
        c.SUBMIT_DATE,            -- When the change request was submitted
        c.IMPLEMENTATION_RESULT,  -- Free-text result field
        c.CLOSURE_CODE,           -- Standardized outcome code (e.g., Successful, Failed)
        w.MANAGER_NAME,           -- Manager's name pulled from the HR employee table
        w.DEPARTMENT              -- Department pulled from the HR employee table
    FROM ITSM.SERVICE_MGMT.CHANGE_REQUESTS c
    -- LEFT JOIN keeps all change records even if the coordinator isn't found in HR data
    LEFT JOIN HR.WORKFORCE.EMPLOYEES w
        ON UPPER(c.COORDINATOR) = UPPER(w.EMPLOYEE_ID)   -- UPPER() on both sides handles mixed-case ID mismatches
    WHERE
        c.SUBMIT_DATE >= DATEADD(DAY, %(lookback)s, CURRENT_DATE())   -- Dynamic rolling window; %(lookback)s = -30
        AND c.CHANGE_TYPE IN ('Standard', 'Normal', 'Emergency')       -- Exclude change types outside scope
        AND c.STATUS NOT IN ('Draft', 'Cancelled')                     -- Only include actionable/completed records
    ORDER BY c.SCHEDULED_START ASC   -- Chronological order makes the output easier to review
"""


def fetch_changes(conn) -> pd.DataFrame:
    """Call run_query() with the change SQL and log the record count."""
    log.info(f"Fetching changes from the last {LOOKBACK_DAYS} days...")
    # Pass the lookback value as a negative integer so DATEADD subtracts days from today
    df = run_query(conn, CHANGE_QUERY, {"lookback": -LOOKBACK_DAYS})
    log.info(f"  Retrieved {len(df):,} change records.")  # :, formats the number with commas (e.g., 1,234)
    return df


# ===========================================================================
# TRANSFORMATION & AUDIT LOGIC
# ===========================================================================

def clean_and_enrich(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize data types, derive calculated fields, and add audit flags.
    Returns a new DataFrame — does not modify the original (defensive copy).
    """
    df = df.copy()   # Work on a copy so the original raw_df passed in is never mutated

    # Convert all timestamp columns from raw strings to proper pandas datetime objects
    # errors="coerce" turns unparseable values into NaT (Not a Time) instead of raising an error
    for col in ["SCHEDULED_START", "SCHEDULED_END", "ACTUAL_START", "ACTUAL_END", "SUBMIT_DATE"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")

    # Calculate planned duration: subtract start from end, convert the timedelta to total seconds, divide to get hours
    df["PLANNED_DURATION_HRS"] = (
        (df["SCHEDULED_END"] - df["SCHEDULED_START"]).dt.total_seconds() / 3600
    ).round(2)   # Round to 2 decimal places for cleaner reporting

    # Same calculation using actual start/end timestamps
    df["ACTUAL_DURATION_HRS"] = (
        (df["ACTUAL_END"] - df["ACTUAL_START"]).dt.total_seconds() / 3600
    ).round(2)

    # Boolean flag: True if the change ran longer than planned (a key compliance signal)
    df["OVERRUN"] = df["ACTUAL_DURATION_HRS"] > df["PLANNED_DURATION_HRS"]

    # Strip leading/trailing whitespace from closure codes and convert to Title Case for consistency
    df["OUTCOME"] = df["CLOSURE_CODE"].str.strip().str.title()

    # Boolean flag: True if the closure code indicates a successful outcome
    df["SUCCESSFUL"] = df["OUTCOME"].isin(["Successful", "Successful With Issues"])

    # .dt.dayofweek returns 0=Monday through 6=Sunday; >= 5 means Saturday or Sunday
    df["WEEKEND_IMPL"] = df["ACTUAL_START"].dt.dayofweek >= 5   # Weekend changes carry higher risk

    return df


def stratified_audit_sample(df: pd.DataFrame, rate: float = AUDIT_SAMPLE_RATE) -> pd.DataFrame:
    """
    Draw a proportional random sample of standard (minor) changes for audit review.

    Stratifying by SUPPORT_GROUP ensures every team is represented in proportion
    to their change volume — the same methodology used in manual audit sampling.
    """
    # Filter to only Standard (minor) changes — these are the ones subject to audit sampling
    minor = df[df["CHANGE_TYPE"] == "Standard"].copy()
    sample_frames = []   # List to collect each group's sample before combining

    # Loop through each unique support group and sample independently
    for group, group_df in minor.groupby("SUPPORT_GROUP"):
        n = max(1, round(len(group_df) * rate))   # Calculate sample size; max(1, ...) ensures at least 1 record per group
        # min(n, len(group_df)) prevents requesting more rows than the group actually has
        # random_state=42 makes the sample reproducible — same seed = same records every run
        sampled = group_df.sample(n=min(n, len(group_df)), random_state=42)
        sample_frames.append(sampled)   # Add this group's sample to the list

    # Combine all group samples into a single DataFrame; reset_index drops the old group-level indexes
    audit_sample = pd.concat(sample_frames, ignore_index=True)
    audit_sample["AUDIT_FLAG"] = True   # Mark every row in the sample so it's identifiable in the output
    log.info(f"  Audit sample size: {len(audit_sample):,} of {len(minor):,} standard changes.")
    return audit_sample


def compute_summary_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate change records into a summary table of KPIs by change type and support group.
    Uses named aggregations (pandas 0.25+) for readable column naming in one step.
    """
    summary = (
        df.groupby(["CHANGE_TYPE", "SUPPORT_GROUP"])   # Group by two dimensions simultaneously
        .agg(
            # Named aggregation syntax: NEW_COL_NAME=("source_column", "aggregation_function")
            TOTAL_CHANGES   = ("CHANGE_ID",           "count"),   # Count of records per group
            SUCCESS_RATE    = ("SUCCESSFUL",           "mean"),    # mean() on a boolean gives proportion (0.0–1.0)
            OVERRUN_RATE    = ("OVERRUN",              "mean"),    # Proportion of changes that ran over time
            AVG_PLANNED_HRS = ("PLANNED_DURATION_HRS", "mean"),   # Average planned change window in hours
            AVG_ACTUAL_HRS  = ("ACTUAL_DURATION_HRS",  "mean"),   # Average actual change duration in hours
            WEEKEND_PCT     = ("WEEKEND_IMPL",         "mean"),    # Proportion implemented on weekends
        )
        .reset_index()   # Converts the groupby index back into regular columns so the DataFrame is flat
    )

    # Convert proportions (0.0–1.0) to percentages (0–100) rounded to 1 decimal place
    summary["SUCCESS_RATE"]  = (summary["SUCCESS_RATE"]  * 100).round(1)
    summary["OVERRUN_RATE"]  = (summary["OVERRUN_RATE"]  * 100).round(1)
    summary["WEEKEND_PCT"]   = (summary["WEEKEND_PCT"]   * 100).round(1)

    # Round hour averages to 2 decimal places for cleaner display
    summary["AVG_PLANNED_HRS"] = summary["AVG_PLANNED_HRS"].round(2)
    summary["AVG_ACTUAL_HRS"]  = summary["AVG_ACTUAL_HRS"].round(2)

    return summary


# ===========================================================================
# EXCEL EXPORT WITH STYLING
# ===========================================================================

# Define reusable style objects at module level so they're created once and shared across all worksheets

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # Dark navy background for header row
ALT_ROW_FILL = PatternFill("solid", fgColor="DCE6F1")   # Light blue for alternating data rows (improves readability)
WARN_FILL    = PatternFill("solid", fgColor="FFEB9C")   # Amber highlight used to flag overrun rows
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)  # White bold text for header cells
BORDER_THIN  = Border(                                   # Thin border applied to every cell for a clean grid look
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)


def style_worksheet(ws, df: pd.DataFrame):
    """
    Apply professional formatting to an openpyxl worksheet:
    - Styled header row with column names
    - Alternating row fill colors
    - Auto-fit column widths based on content length
    - Frozen top row and auto-filter dropdowns
    """
    # --- Header row ---
    for col_idx, col_name in enumerate(df.columns, start=1):   # enumerate starts at 1 to match Excel's 1-based columns
        # Replace underscores with spaces and title-case the column name for a readable header
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.fill      = HEADER_FILL    # Apply navy background
        cell.font      = HEADER_FONT    # Apply white bold font
        cell.border    = BORDER_THIN    # Apply cell border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)   # Center and wrap long headers

    # --- Data rows ---
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):   # start=2 skips the header row
        # Apply alternating fill: even rows get light blue, odd rows get no fill (default white)
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER_THIN
            cell.fill   = fill

    # --- Auto-fit column widths ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        # Find the longest value in the column (or the column name itself, whichever is wider)
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max()   # .iloc uses 0-based indexing, so subtract 1
        )
        # get_column_letter(3) → "C"; cap width at 40 to prevent excessively wide columns
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes    = "A2"           # Freeze the header row so it stays visible when scrolling down
    ws.auto_filter.ref = ws.dimensions  # Apply Excel auto-filter dropdowns across all columns


def export_to_excel(detail_df: pd.DataFrame, summary_df: pd.DataFrame,
                    audit_df: pd.DataFrame) -> str:
    """
    Write three DataFrames to separate sheets in a single .xlsx file,
    then reopen the file with openpyxl to apply visual styling.
    Returns the file path of the saved report.
    """
    # Build a timestamped filename so each run produces a unique file (no overwriting)
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"change_audit_report_{timestamp}.xlsx")

    # pandas ExcelWriter handles creating the .xlsx and writing each DataFrame to its own sheet
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary Metrics", index=False)  # index=False omits the DataFrame row numbers
        detail_df.to_excel(writer, sheet_name="All Changes",     index=False)
        audit_df.to_excel(writer, sheet_name="Audit Sample",    index=False)
    # The 'with' block auto-saves and closes the file here

    # Reopen the saved file with openpyxl so we can apply styling (pandas ExcelWriter doesn't support this directly)
    wb = load_workbook(output_path)
    style_worksheet(wb["Summary Metrics"], summary_df)   # Style each sheet using the helper function above
    style_worksheet(wb["All Changes"],     detail_df)
    style_worksheet(wb["Audit Sample"],    audit_df)

    # --- Highlight overrun rows in the detail sheet ---
    ws = wb["All Changes"]
    # Find the 1-based column position of "OVERRUN" so we can check its value row by row
    overrun_col = list(detail_df.columns).index("OVERRUN") + 1   # .index() is 0-based, so +1 for Excel
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):       # Skip row 1 (header)
        if row[overrun_col - 1].value is True:    # row[] is 0-based, so -1 to align with the column position
            for cell in row:
                cell.fill = WARN_FILL   # Paint the entire row amber to draw attention to the overrun

    wb.save(output_path)   # Save the styled workbook back to the same path
    log.info(f"  Report saved: {output_path}")
    return output_path     # Return the path so the email function knows what to attach


# ===========================================================================
# EMAIL NOTIFICATION
# ===========================================================================

def build_html_body(summary_df: pd.DataFrame, total: int, sample_size: int) -> str:
    """
    Build an HTML email body containing a run summary and an inline metrics table.
    pandas' .to_html() renders the DataFrame as a styled HTML <table>.
    """
    run_date   = datetime.now().strftime("%B %d, %Y")   # e.g., "June 01, 2025"
    # Convert the summary DataFrame to an HTML table string; border=0 removes default HTML borders (CSS handles it)
    table_html = summary_df.to_html(index=False, border=0, classes="metrics-table")

    # f-string with embedded CSS and the table HTML; double braces {{ }} are needed inside f-strings to produce literal { }
    return f"""
    <html><head>
    <style>
      body {{ font-family: Calibri, Arial, sans-serif; font-size: 13px; }}
      .metrics-table {{ border-collapse: collapse; width: 100%; }}
      .metrics-table th {{ background-color: #1F3864; color: white; padding: 6px 10px; }}
      .metrics-table td {{ border: 1px solid #ccc; padding: 5px 10px; }}
      .metrics-table tr:nth-child(even) {{ background-color: #DCE6F1; }}
    </style>
    </head><body>
    <p>Team,</p>
    <p>The automated change audit pipeline has completed for the period ending <strong>{run_date}</strong>.</p>
    <ul>
      <li><strong>Total changes reviewed:</strong> {total:,}</li>
      <li><strong>Audit sample (standard changes):</strong> {sample_size:,}</li>
      <li><strong>Lookback window:</strong> {LOOKBACK_DAYS} days</li>
    </ul>
    <p>Summary metrics by change type and support group:</p>
    {table_html}
    <p>Full detail and audit sample attached. Please review flagged overruns (highlighted amber).</p>
    <br>
    <p>— Change Management Automation</p>
    </body></html>
    """


def send_email(subject: str, html_body: str, attachment_path: str):
    """
    Construct a MIME email with an HTML body and an Excel file attachment,
    then send it via SMTP with STARTTLS encryption.
    """
    msg            = MIMEMultipart("alternative")       # "alternative" allows the message to hold multiple content parts
    msg["Subject"] = subject
    msg["From"]    = SENDER
    msg["To"]      = ", ".join(RECIPIENTS)              # Email headers expect a comma-separated string, not a list
    msg.attach(MIMEText(html_body, "html"))             # Attach the HTML body as the email's readable content

    # --- Attach the Excel file ---
    with open(attachment_path, "rb") as f:              # "rb" = read binary; Excel files are binary, not text
        part = MIMEBase("application", "octet-stream")  # "octet-stream" is the generic MIME type for binary files
        part.set_payload(f.read())                      # Load the raw file bytes into the MIME part
    encoders.encode_base64(part)                        # Encode bytes as base64 so they survive email transmission safely
    # Tell the email client this is a downloadable file attachment and what to name it
    part.add_header("Content-Disposition",
                    f"attachment; filename={os.path.basename(attachment_path)}")
    msg.attach(part)   # Add the attachment part to the email message object

    # --- Send via SMTP ---
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:  # Opens a connection; 'with' auto-closes it when done
        server.starttls()                                # Upgrade the connection to encrypted TLS before sending credentials
        server.login(SENDER, os.getenv("EMAIL_PASSWORD"))   # Authenticate with the mail server
        server.sendmail(SENDER, RECIPIENTS, msg.as_string())  # RECIPIENTS must be a list here (vs. the header string above)
    log.info(f"  Email sent to: {', '.join(RECIPIENTS)}")


# ===========================================================================
# MAIN PIPELINE — orchestrates all steps in order
# ===========================================================================

def run_pipeline():
    """
    Entry point for the full pipeline:
      1. Connect to Snowflake
      2. Extract change records
      3. Clean and enrich the data
      4. Generate stratified audit sample
      5. Compute summary KPIs
      6. Export to Excel
      7. Send email notification
    """
    log.info("=== Change Audit Pipeline Starting ===")
    conn = get_connection()   # Open Snowflake connection once and reuse it for all queries

    try:
        # Step 1 — Extract raw data from Snowflake
        raw_df = fetch_changes(conn)

        # Step 2 — Clean and derive calculated columns
        log.info("Cleaning and enriching change data...")
        clean_df = clean_and_enrich(raw_df)

        # Step 3 — Pull the stratified audit sample from standard changes only
        log.info("Generating stratified audit sample...")
        audit_df = stratified_audit_sample(clean_df)

        # Step 4 — Aggregate into a KPI summary table
        log.info("Computing summary metrics...")
        summary_df = compute_summary_metrics(clean_df)

        # Step 5 — Write all three DataFrames to a styled Excel workbook
        log.info("Exporting to Excel...")
        report_path = export_to_excel(clean_df, summary_df, audit_df)

        # Step 6 — Build the email and send with the report attached
        log.info("Sending email notification...")
        subject = (
            f"Change Audit Report — {datetime.now().strftime('%B %Y')} "
            f"({len(clean_df):,} Changes)"
        )
        html = build_html_body(summary_df, len(clean_df), len(audit_df))
        send_email(subject, html, report_path)

    finally:
        # Always close the Snowflake connection, even if an error occurred mid-pipeline
        conn.close()
        log.info("Snowflake connection closed.")

    log.info("=== Pipeline Complete ===")


# Only run the pipeline when this script is executed directly (not when imported as a module)
if __name__ == "__main__":
    run_pipeline()
